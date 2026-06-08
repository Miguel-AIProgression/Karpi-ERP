"""Importeer productie-only orders uit totaalplanning_cleaned_v2.xlsx naar RugFlow.

Dry-run default; --commit roept de RPC import_productie_only_order aan.
Groepeert regels per Basta-ordernr; rapporteert niet-herkende afwerkingscodes.
"""
from __future__ import annotations
import argparse, datetime as dt, os, sys
from collections import defaultdict
from openpyxl import load_workbook
from lib import snijlijst_parser as P
from lib.afwerking_mapper import map_afwerking_code, is_herkend

BESTAND = "totaalplanning_cleaned_v2.xlsx"
SHEET   = "Snijden Karpi op kwal"


def verzendweek_naar_datum(s) -> dt.date | None:
    """'24-2026' -> maandag van ISO-week 24 in 2026. Ongeldig/leeg -> None."""
    s = str(s).strip()
    if "-" not in s:
        return None
    wk, jaar = s.split("-", 1)
    try:
        return dt.date.fromisocalendar(int(jaar), int(wk), 1)  # maandag
    except ValueError:
        return None


def bepaal_vorm(maat1, maat2, omschrijving) -> str:
    """Leid de vorm af: RND-maat2 -> rond, 'OVAAL' in omschrijving -> ovaal, anders rechthoek."""
    m2 = str(maat2).strip().upper()
    if m2 == "RND":
        return "rond"
    if "OVAAL" in str(omschrijving).upper():
        return "ovaal"
    return "rechthoek"


def rij_naar_regel(rij) -> dict | None:
    rij = list(rij) + [""] * (max(P.PL_OPMERKING, P.PL_RGL) + 1 - len(rij))
    ordernr = P.normaliseer_key(rij[P.PL_ORDERNR])
    if ordernr is None:
        return None
    kk = P.parse_artikelcode_kwal_kleur(rij[P.PL_ARTIKELCODE]) or ("", "")
    grof = rij[14]   # GROF-afwerking (geen constante in snijlijst_parser)
    fijn = rij[6]    # FIJN-afwerking (geen constante in snijlijst_parser)
    omschr = P._norm(rij[1])
    try:
        aantal = int(float(P._norm(rij[P.PL_AANTAL]))) if P._norm(rij[P.PL_AANTAL]) else 1
    except ValueError:
        aantal = 1
    # maatwerk_lengte_cm/breedte_cm = de RUWE maat1/maat2 (de fysieke snijmaat),
    # NIET P.breedte_lengte_uit_maten (die normaliseert naar max/min voor
    # rol-allocatie). rechthoek: lengte=int(maat1), breedte=int(maat2).
    # RND (maat2 upper == 'RND'): beide = diameter = int(maat1).
    # Niet-parsebare maten -> return None (skip rij, net als parse_planning_rij).
    m1_raw = P._norm(rij[P.PL_MAAT1])
    m2_raw = P._norm(rij[P.PL_MAAT2])
    try:
        if m2_raw.upper() == "RND":
            lengte_cm = breedte_cm = int(float(m1_raw))
        else:
            lengte_cm = int(float(m1_raw))
            breedte_cm = int(float(m2_raw))
    except ValueError:
        return None
    return {
        "oud_order_nr": int(ordernr),
        "debiteur_nr": P.normaliseer_key(rij[11]),
        "debiteur_naam": P._norm(rij[12]),
        "regelnummer": int(P.normaliseer_key(rij[P.PL_RGL]) or "1"),
        "omschrijving": omschr or "Maatwerk",
        "orderaantal": max(aantal, 1),
        "maatwerk_kwaliteit_code": kk[0],
        "maatwerk_kleur_code": kk[1],
        "maatwerk_lengte_cm": lengte_cm,
        "maatwerk_breedte_cm": breedte_cm,
        "maatwerk_afwerking": map_afwerking_code(grof, fijn, omschr),
        "afwerking_herkend": is_herkend(grof, fijn),
        "maatwerk_vorm": bepaal_vorm(rij[P.PL_MAAT1], rij[P.PL_MAAT2], omschr),
        "snijden_uit_standaardmaat": P.is_snijden_uit(rij[P.PL_OPMERKING]),
        "maatwerk_instructies": P._norm(rij[P.PL_OPMERKING]),
        "afleverdatum": verzendweek_naar_datum(rij[17]),
    }


def lees_regels(pad):
    wb = load_workbook(pad, read_only=True, data_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[2:]:               # data vanaf idx 2 (header idx 1)
        regel = rij_naar_regel(r)
        if regel:
            out.append(regel)
    wb.close()
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true")
    ap.add_argument("--bestand", default=BESTAND)
    args = ap.parse_args()

    regels = lees_regels(args.bestand)
    per_order = defaultdict(list)
    for r in regels:
        per_order[r["oud_order_nr"]].append(r)

    onherkend = [r for r in regels if not r["afwerking_herkend"]]
    uit_std   = [r for r in regels if r["snijden_uit_standaardmaat"]]
    print(f"Regels: {len(regels)} | Orders: {len(per_order)} | "
          f"uit-standaardmaat: {len(uit_std)} | afwerking-default-gebruikt: {len(onherkend)}")
    if onherkend:
        print("  Niet-herkende afwerking (krijgt 'B' default) - controleer:")
        for r in onherkend[:40]:
            print(f"    Basta {r['oud_order_nr']} rgl {r['regelnummer']}: "
                  f"GROF/FIJN onbekend -> B  ({r['omschrijving']})")

    if not args.commit:
        print("DRY-RUN - niets weggeschreven. Draai met --commit om te importeren.")
        return

    from supabase import create_client
    sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_SERVICE_KEY"])
    nieuw = bestaand = 0
    for oud_nr, rs in per_order.items():
        header = {
            "oud_order_nr": oud_nr,
            "debiteur_nr": rs[0]["debiteur_nr"],
            "debiteur_naam": rs[0]["debiteur_naam"],
            "afleverdatum": rs[0]["afleverdatum"].isoformat() if rs[0]["afleverdatum"] else None,
        }
        payload_regels = [{k: v for k, v in r.items()
                           if k not in ("debiteur_nr", "debiteur_naam", "afleverdatum", "afwerking_herkend", "oud_order_nr")}
                          for r in rs]
        res = sb.rpc("import_productie_only_order",
                     {"p_header": header, "p_regels": payload_regels}).execute()
        row = res.data[0] if res.data else {}
        if row.get("was_existing"):
            bestaand += 1
        else:
            nieuw += 1
    print(f"Klaar: {nieuw} nieuw, {bestaand} bestaand (idempotent overgeslagen).")


if __name__ == "__main__":
    main()
