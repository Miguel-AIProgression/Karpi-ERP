"""
RugFlow ERP - aanvullende prijslijsten import.

Leest prijslijsten.zip + toevoegingprijslijsten.zip via een manifest en zet:
- prijslijst_headers
- prijslijst_regels
- debiteuren.prijslijst_nr

Default is dry-run. Gebruik --apply voor echte Supabase-mutaties.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from supabase import create_client

from config import BASE_DIR, DEBITEUREN_FILE, SUPABASE_KEY, SUPABASE_URL

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

MANIFEST_PATH = BASE_DIR / "import" / "prijslijsten_aanvulling_manifest.json"
REPORT_DIR = BASE_DIR / "import" / "rapporten"
BATCH_SIZE = 500


def zpad(value: object) -> str:
    return str(value).strip().zfill(4)


def parse_date(value: str | None) -> str | None:
    if not value:
        return None
    match = re.search(r"(\d{1,2})[.\-](\d{1,2})[.\-](\d{2,4})", value)
    if not match:
        return None
    day, month, year = match.groups()
    if len(year) == 2:
        year = "20" + year
    try:
        return datetime(int(year), int(month), int(day)).strftime("%Y-%m-%d")
    except ValueError:
        return None


def decimal_or_none(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.replace(",", ".").strip()
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load_manifest() -> dict:
    return json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))


def load_source_debiteur_mapping() -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(DEBITEUREN_FILE, read_only=True, data_only=True)
    ws = wb["Debiteuren"]
    header = next(ws.iter_rows(values_only=True))
    idx = {name: i for i, name in enumerate(header)}
    mapping: dict[str, list[dict]] = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_prijslijst = row[idx["Prijslijst"]]
        status = str(row[idx["Status"]] or "").strip()
        if not raw_prijslijst or status.lower() != "actief":
            continue
        match = re.match(r"\s*(\d+)", str(raw_prijslijst))
        if not match:
            continue
        nr = zpad(match.group(1))
        mapping[nr].append(
            {
                "debiteur_nr": int(row[idx["Debiteur"]]),
                "naam": str(row[idx["Naam"]] or "").strip(),
                "plaats": str(row[idx["Plaats"]] or "").strip(),
                "bron_prijslijst": str(raw_prijslijst).strip(),
            }
        )

    wb.close()
    return dict(mapping)


def read_zip_member(source_zip: Path, filename: str) -> bytes:
    with zipfile.ZipFile(source_zip, "r") as zf:
        candidates = [
            name
            for name in zf.namelist()
            if Path(name).name == filename and "__MACOSX" not in name and not Path(name).name.startswith("~$")
        ]
        if not candidates:
            raise FileNotFoundError(f"{filename} niet gevonden in {source_zip.name}")
        return zf.read(candidates[0])


def read_price_list(xlsx_bytes: bytes, filename: str, expected_nr: str) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 4:
        raise ValueError(f"{filename} bevat te weinig rijen")

    meta = rows[1]
    excel_nr = zpad(int(meta[0])) if meta[0] is not None else None
    if excel_nr and excel_nr != expected_nr:
        raise ValueError(f"{filename}: manifest zegt {expected_nr}, Excel zegt {excel_nr}")

    naam = str(meta[1]).strip() if len(meta) > 1 and meta[1] else filename.replace(".xlsx", "")
    geldig_vanaf = parse_date(naam) or parse_date(filename)
    regels = []

    for row in rows[3:]:
        if not row or row[0] is None:
            continue
        try:
            artikelnr = str(int(row[0]))
        except (TypeError, ValueError):
            continue

        prijs = decimal_or_none(row[3] if len(row) > 3 else None)
        gewicht = decimal_or_none(row[5] if len(row) > 5 else None)
        regels.append(
            {
                "artikelnr": artikelnr,
                "omschrijving": str(row[1]).strip() if len(row) > 1 and row[1] else None,
                "omschrijving_2": str(row[2]).strip() if len(row) > 2 and row[2] else None,
                "prijs": prijs if prijs is not None else 0.0,
                "gewicht": gewicht,
                "ean_code": None,
            }
        )

    return {
        "nr": expected_nr,
        "naam": naam,
        "geldig_vanaf": geldig_vanaf,
        "regels": regels,
    }


def fetch_all(sb, table: str, select_cols: str) -> list[dict]:
    all_rows: list[dict] = []
    offset = 0
    while True:
        result = sb.table(table).select(select_cols).range(offset, offset + 999).execute()
        rows = result.data or []
        all_rows.extend(rows)
        if len(rows) < 1000:
            return all_rows
        offset += 1000


def upsert_batch(sb, table: str, records: list[dict], on_conflict: str | None = None) -> None:
    for i in range(0, len(records), BATCH_SIZE):
        kwargs = {"on_conflict": on_conflict} if on_conflict else {}
        sb.table(table).upsert(records[i : i + BATCH_SIZE], **kwargs).execute()


def classify_product(omschrijving: str | None) -> str:
    text = (omschrijving or "").upper()
    if "BREED" in text:
        return "rol"
    if "CA:" in text:
        match = re.search(r"CA:\s*(\d+)\s*[xX]\s*(\d+)", text)
        if match and int(match.group(1)) * int(match.group(2)) < 10000:
            return "staaltje"
        return "vast"
    return "overig"


def build_import_set(manifest: dict, source_map: dict[str, list[dict]]) -> tuple[list[dict], list[str]]:
    entries = []
    warnings = []

    for item in manifest["files"]:
        prijslijst_nr = zpad(item["prijslijst_nr"])
        debiteuren = source_map.get(prijslijst_nr, [])
        expected = sorted(item.get("expected_debiteur_nrs") or [])
        actual = sorted(d["debiteur_nr"] for d in debiteuren)
        if expected and actual != expected:
            warnings.append(
                f"{item['filename']}: expected_debiteur_nrs {expected} wijkt af van debiteuren-export {actual}"
            )
        if not debiteuren:
            warnings.append(f"{item['filename']}: geen actieve debiteuren gevonden voor prijslijst {prijslijst_nr}")

        source_zip = BASE_DIR / item["source_zip"]
        xlsx_bytes = read_zip_member(source_zip, item["filename"])
        parsed = read_price_list(xlsx_bytes, item["filename"], prijslijst_nr)
        parsed.update(
            {
                "source_zip": item["source_zip"],
                "filename": item["filename"],
                "sha256": hashlib.sha256(xlsx_bytes).hexdigest(),
                "debiteuren": debiteuren,
                "match_basis": item.get("match_basis"),
            }
        )
        entries.append(parsed)

    return entries, warnings


def validate(entries: list[dict], sb) -> dict:
    known_products = {row["artikelnr"] for row in fetch_all(sb, "producten", "artikelnr")}
    known_debiteuren = {
        int(row["debiteur_nr"]): row
        for row in fetch_all(sb, "debiteuren", "debiteur_nr, naam, prijslijst_nr, status")
    }
    existing_counts = Counter(
        row["prijslijst_nr"] for row in fetch_all(sb, "prijslijst_regels", "prijslijst_nr")
    )

    headers = []
    regels = []
    links = []
    missing_products = {}
    problems = []

    seen_headers = set()
    seen_links = set()
    seen_regel_keys = set()

    for entry in entries:
        nr = entry["nr"]
        if nr in seen_headers:
            problems.append(f"Dubbele prijslijst in manifest: {nr}")
        seen_headers.add(nr)

        headers.append(
            {
                "nr": nr,
                "naam": entry["naam"],
                "geldig_vanaf": entry["geldig_vanaf"],
                "actief": True,
            }
        )

        for row in entry["regels"]:
            key = (nr, row["artikelnr"])
            if key in seen_regel_keys:
                problems.append(f"Dubbele regel binnen batch: {nr}/{row['artikelnr']}")
                continue
            seen_regel_keys.add(key)
            if row["artikelnr"] not in known_products and row["artikelnr"] not in missing_products:
                missing_products[row["artikelnr"]] = {
                    "artikelnr": row["artikelnr"],
                    "omschrijving": row["omschrijving"] or "Onbekend product",
                    "verkoopprijs": row["prijs"],
                    "gewicht_kg": row["gewicht"],
                    "voorraad": 0,
                    "gereserveerd": 0,
                    "vrije_voorraad": 0,
                    "product_type": classify_product(row["omschrijving"]),
                    "actief": True,
                }
            regels.append({"prijslijst_nr": nr, **row})

        for debiteur in entry["debiteuren"]:
            debiteur_nr = debiteur["debiteur_nr"]
            if debiteur_nr not in known_debiteuren:
                problems.append(f"Debiteur {debiteur_nr} bestaat niet in Supabase")
                continue
            link_key = (debiteur_nr, nr)
            if link_key in seen_links:
                continue
            seen_links.add(link_key)
            links.append(
                {
                    "debiteur_nr": debiteur_nr,
                    "naam": known_debiteuren[debiteur_nr]["naam"],
                    "oude_prijslijst_nr": known_debiteuren[debiteur_nr].get("prijslijst_nr"),
                    "nieuwe_prijslijst_nr": nr,
                }
            )

    return {
        "headers": headers,
        "regels": regels,
        "links": links,
        "missing_products": list(missing_products.values()),
        "problems": problems,
        "existing_counts": existing_counts,
    }


def write_report(manifest: dict, entries: list[dict], validation: dict, warnings: list[str], applied: bool) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    path = REPORT_DIR / f"prijslijsten-aanvulling-{stamp}.md"

    lines = [
        f"# Prijslijsten aanvulling - {'apply' if applied else 'dry-run'}",
        "",
        f"Batch: `{manifest['batch']}`",
        f"Moment: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "## Samenvatting",
        "",
        f"- Bestanden: {len(entries)}",
        f"- Prijslijsten: {len(validation['headers'])}",
        f"- Regels: {len(validation['regels'])}",
        f"- Debiteurkoppelingen: {len(validation['links'])}",
        f"- Nieuw aan te maken producten: {len(validation['missing_products'])}",
        f"- Waarschuwingen: {len(warnings)}",
        f"- Blokkerende problemen: {len(validation['problems'])}",
        "",
        "## Bestanden",
        "",
    ]

    for entry in entries:
        current_count = validation["existing_counts"].get(entry["nr"], 0)
        lines.extend(
            [
                f"### {entry['nr']} - {entry['naam']}",
                f"- Bestand: `{entry['source_zip']}::{entry['filename']}`",
                f"- Geldig vanaf: {entry['geldig_vanaf'] or 'onbekend'}",
                f"- Excel-regels: {len(entry['regels'])}",
                f"- Bestaande Supabase-regels voor import: {current_count}",
                f"- Debiteuren: {len(entry['debiteuren'])}",
                f"- SHA256: `{entry['sha256']}`",
                "",
            ]
        )
        for deb in entry["debiteuren"][:12]:
            lines.append(f"  - {deb['debiteur_nr']} - {deb['naam']} ({deb['plaats']})")
        if len(entry["debiteuren"]) > 12:
            lines.append(f"  - ... en {len(entry['debiteuren']) - 12} meer")
        lines.append("")

    if warnings:
        lines.extend(["## Waarschuwingen", ""])
        lines.extend(f"- {warning}" for warning in warnings)
        lines.append("")

    if validation["problems"]:
        lines.extend(["## Blokkerende problemen", ""])
        lines.extend(f"- {problem}" for problem in validation["problems"])
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def apply_import(sb, validation: dict) -> None:
    if validation["missing_products"]:
        upsert_batch(sb, "producten", validation["missing_products"], on_conflict="artikelnr")
    upsert_batch(sb, "prijslijst_headers", validation["headers"], on_conflict="nr")
    upsert_batch(sb, "prijslijst_regels", validation["regels"], on_conflict="prijslijst_nr,artikelnr")
    for link in validation["links"]:
        sb.table("debiteuren").update({"prijslijst_nr": link["nieuwe_prijslijst_nr"]}).eq(
            "debiteur_nr", link["debiteur_nr"]
        ).execute()


def main() -> int:
    parser = argparse.ArgumentParser(description="Importeer aanvullende prijslijsten.")
    parser.add_argument("--apply", action="store_true", help="Voer Supabase-mutaties uit.")
    args = parser.parse_args()

    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERROR: Supabase URL/key ontbreekt. Check import/.env")
        return 1

    manifest = load_manifest()
    source_map = load_source_debiteur_mapping()
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    entries, warnings = build_import_set(manifest, source_map)
    validation = validate(entries, sb)
    report_path = write_report(manifest, entries, validation, warnings, applied=args.apply)

    print(f"Bestanden: {len(entries)}")
    print(f"Prijslijsten: {len(validation['headers'])}")
    print(f"Regels: {len(validation['regels'])}")
    print(f"Debiteurkoppelingen: {len(validation['links'])}")
    print(f"Nieuw aan te maken producten: {len(validation['missing_products'])}")
    print(f"Waarschuwingen: {len(warnings)}")
    print(f"Blokkerende problemen: {len(validation['problems'])}")
    print(f"Rapport: {report_path}")

    if validation["problems"]:
        print("ERROR: import afgebroken door blokkerende problemen.")
        return 1

    if not args.apply:
        print("Dry-run klaar. Gebruik --apply om te importeren.")
        return 0

    apply_import(sb, validation)
    print("Import uitgevoerd.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
