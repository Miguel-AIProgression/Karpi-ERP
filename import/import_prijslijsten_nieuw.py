"""
RugFlow ERP — Nieuwe prijslijsten importeren + klanten koppelen (2026-05-29)

Verwerkt: prijslijsten_nieuw.zip (10 Excel-bestanden)
Brondata klanten: Desktop/binnendienst/debadres_alles-4.xlsx

Stap 1: prijslijst_headers aanmaken
Stap 2: prijslijst_regels importeren (artikelprijzen)
Stap 3: debiteuren.prijslijst_nr koppelen

Gebruik:
  python import_prijslijsten_nieuw.py           # dry-run
  python import_prijslijsten_nieuw.py --apply   # echt uitvoeren
"""
from __future__ import annotations

import argparse
import io
import re
import sys
import zipfile
from collections import defaultdict
from datetime import date
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from supabase import create_client

from config import BASE_DIR, SUPABASE_KEY, SUPABASE_URL

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: Supabase URL/Key niet gevonden. Check import/.env")
    sys.exit(1)

sb = create_client(SUPABASE_URL, SUPABASE_KEY)

ZIP_PATH = BASE_DIR / "prijslijsten_nieuw.zip"
DEBITEUREN_FILE = Path.home() / "Desktop" / "binnendienst" / "debadres_alles-4.xlsx"
TARGET_NRS = {"143", "160", "161", "180", "181", "185", "195", "197", "206", "250"}
BATCH = 500


def zpad(nr: str | int) -> str:
    return str(nr).strip().zfill(4)


def parse_datum(tekst: str) -> str | None:
    m = re.search(r"(\d{1,2})[.\-](\d{1,2})[.\-](\d{4})", tekst)
    if not m:
        return None
    d, mo, y = m.groups()
    try:
        return date(int(y), int(mo), int(d)).isoformat()
    except ValueError:
        return None


def decimal(val: object) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


# ── Excel-bestand inlezen ──────────────────────────────────────────
def parse_excel(data: bytes, prijslijst_nr: str) -> tuple[dict, list[dict]]:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    header: dict = {}
    regels: list[dict] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # "Prijslijst-overzicht" titel-rij
        if i == 1:
            naam = str(row[2]).strip() if row[2] else f"Prijslijst {prijslijst_nr}"
            header = {
                "nr": prijslijst_nr,
                "naam": naam,
                "geldig_vanaf": parse_datum(naam),
                "actief": True,
            }
            continue
        if i == 2:
            continue  # kolomkoppen

        artikelnr = row[0]
        prijs_raw = row[4]

        if not artikelnr or prijs_raw is None:
            continue

        prijs = decimal(prijs_raw)
        if prijs is None:
            continue

        omschrijving = str(row[2]).strip() if row[2] else None
        omschrijving_2 = str(row[3]).strip() if row[3] else None
        gewicht = decimal(row[6]) if len(row) > 6 else None

        regels.append({
            "prijslijst_nr": prijslijst_nr,
            "artikelnr": str(int(artikelnr)).zfill(9),
            "omschrijving": omschrijving,
            "omschrijving_2": omschrijving_2,
            "prijs": prijs,
            "gewicht": gewicht,
        })

    return header, regels


# ── Klantenkoppeling uit debadres_alles-4.xlsx ────────────────────
def lees_klanten() -> tuple[dict[str, list[int]], dict[str, str]]:
    wb = openpyxl.load_workbook(DEBITEUREN_FILE, read_only=True, data_only=True)
    ws = wb.active

    debiteur_col = prijslijst_col = None
    klanten: dict[str, list[int]] = defaultdict(list)
    pl_namen: dict[str, str] = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if i == 1:
            for j, v in enumerate(row):
                if v == "Debiteur":   debiteur_col = j
                if v == "Prijslijst": prijslijst_col = j
            continue
        debnr = row[debiteur_col] if debiteur_col is not None else None
        prijs = row[prijslijst_col] if prijslijst_col is not None else None
        if not prijs or not debnr:
            continue
        m = re.match(r"0*(\d+)\s*-?\s*(.*)", str(prijs).strip())
        if not m:
            continue
        nr, naam = m.group(1), m.group(2).strip()
        if nr not in TARGET_NRS:
            continue
        if nr not in pl_namen and naam:
            pl_namen[nr] = naam
        try:
            klanten[nr].append(int(debnr))
        except (TypeError, ValueError):
            pass

    return klanten, pl_namen


def fetch_bekende_artikelnrs() -> set[str]:
    """Haal alle artikelnrs op uit producten (gepagineerd)."""
    bekend: set[str] = set()
    offset = 0
    while True:
        res = sb.table("producten").select("artikelnr") \
            .range(offset, offset + 999).execute()
        for r in res.data:
            bekend.add(r["artikelnr"])
        if len(res.data) < 1000:
            break
        offset += 1000
    return bekend


def upsert_batch(table: str, records: list[dict], on_conflict: str) -> None:
    for i in range(0, len(records), BATCH):
        sb.table(table).upsert(records[i:i+BATCH], on_conflict=on_conflict).execute()


def fetch_huidige(debnrs: list[int]) -> dict[int, str | None]:
    result: dict[int, str | None] = {}
    for i in range(0, len(debnrs), 500):
        res = sb.table("debiteuren").select("debiteur_nr,prijslijst_nr") \
            .in_("debiteur_nr", debnrs[i:i+500]).execute()
        for r in res.data:
            result[r["debiteur_nr"]] = r["prijslijst_nr"]
    return result


# ── Main ──────────────────────────────────────────────────────────
def main(apply: bool) -> None:
    mode = "APPLY" if apply else "DRY-RUN"
    print(f"=== Prijslijsten import [{mode}] ===\n")

    if not ZIP_PATH.exists():
        print(f"ERROR: {ZIP_PATH} niet gevonden")
        sys.exit(1)

    # ── Stap 1 + 2: headers + regels ───────────────────────────────
    print("── Stap 1+2: prijslijst_headers + prijslijst_regels ──")
    print("  Bekende artikelnrs ophalen uit producten-tabel...")
    bekende_artikelnrs = fetch_bekende_artikelnrs()
    print(f"  {len(bekende_artikelnrs)} artikelnrs bekend in DB\n")

    alle_headers = []
    totaal_regels = totaal_skip = 0

    with zipfile.ZipFile(ZIP_PATH) as zf:
        for filename in sorted(zf.namelist()):
            if filename.startswith("__MACOSX") or filename.startswith("._"):
                continue
            if not filename.endswith(".xlsx"):
                continue

            m = re.search(r"0*(\d+)", Path(filename).stem)
            if not m or m.group(1) not in TARGET_NRS:
                print(f"  SKIP {filename} — nummer niet herkend")
                continue

            nr = zpad(m.group(1))
            data = zf.read(filename)
            header, regels = parse_excel(data, nr)

            # Filter artikelnrs die niet in producten staan
            bekende = [r for r in regels if r["artikelnr"] in bekende_artikelnrs]
            onbekend = [r for r in regels if r["artikelnr"] not in bekende_artikelnrs]

            print(f"  {nr}  '{header['naam']}'")
            print(f"       {len(bekende)} regels importeren, {len(onbekend)} overgeslagen (artikelnr niet in DB)")
            if onbekend:
                for r in onbekend[:3]:
                    print(f"       skip: {r['artikelnr']}  {r['omschrijving']}")
                if len(onbekend) > 3:
                    print(f"       ... en {len(onbekend)-3} meer")

            alle_headers.append(header)
            totaal_regels += len(bekende)
            totaal_skip   += len(onbekend)

            if apply:
                sb.table("prijslijst_headers").upsert(header, on_conflict="nr").execute()
                if bekende:
                    upsert_batch("prijslijst_regels", bekende, "prijslijst_nr,artikelnr")

    if apply:
        print(f"\n  → {len(alle_headers)} headers + {totaal_regels} regels opgeslagen, {totaal_skip} overgeslagen")
    else:
        print(f"\n  → [dry-run] zou {len(alle_headers)} headers + {totaal_regels} regels opslaan, {totaal_skip} overgeslagen")

    # ── Stap 3: klanten koppelen ────────────────────────────────────
    print("\n── Stap 3: klanten koppelen ──")
    klanten_per_pl, _ = lees_klanten()

    alle_debnrs = sorted(set(d for lst in klanten_per_pl.values() for d in lst))
    huidig = fetch_huidige(alle_debnrs)

    grand_update = grand_al_goed = grand_skip = 0

    for nr in sorted(TARGET_NRS, key=int):
        padded = zpad(nr)
        debnrs = klanten_per_pl.get(nr, [])
        if not debnrs:
            continue

        in_db      = [d for d in debnrs if d in huidig]
        al_goed    = [d for d in in_db  if huidig[d] == padded]
        te_updaten = [d for d in in_db  if huidig[d] != padded]
        skip       = [d for d in debnrs if d not in huidig]

        print(f"\n  {padded}: {len(te_updaten)} koppelen, {len(al_goed)} al correct, {len(skip)} niet in DB")

        if apply and te_updaten:
            for i in range(0, len(te_updaten), 200):
                chunk = te_updaten[i:i+200]
                sb.table("debiteuren").update({"prijslijst_nr": padded}) \
                    .in_("debiteur_nr", chunk).execute()

        grand_update  += len(te_updaten)
        grand_al_goed += len(al_goed)
        grand_skip    += len(skip)

    print(f"\n=== Samenvatting [{mode}] ===")
    print(f"  Prijslijst regels:  {totaal_regels}")
    print(f"  Klanten gekoppeld:  {grand_update}")
    print(f"  Al correct:         {grand_al_goed}")
    print(f"  Niet in DB (skip):  {grand_skip}")
    if not apply:
        print("\n  Gebruik --apply om echt uit te voeren.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    main(apply=args.apply)
