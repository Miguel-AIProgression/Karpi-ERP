"""
Voorraad-update vaste maten — GENERIEK (elke periodieke vrije-voorraadlijst)
============================================================================
Werkt de stuks-voorraad bij uit een Karpi-export "Ovz. vrije voorraad —
alle artikelen" (`Voorraadlijst <datum>.xls`). Vervangt de losse gedateerde
scripts (update_voorraad_2026_05.py / _2026_06_01.py): geef het bestandspad
gewoon als argument mee.

Beslissingen (vastgelegd met Karpi):
  - Scope: ALLEEN product_type='vast'. Staaltje/rol/overig NIET aangeraakt.
  - Sleutel: kolom A 'Artikelnr' -> producten.artikelnr (PK).
  - Waarde:  kolom H 'Vrije voorraad' (= fysiek D - oude reserveringen F).
    HERZIEN 2026-06-15 (was kolom D 'Voorraad' (FYSIEK) sinds 2026-06-08).
    Reden: het oude systeem houdt de actuele voorraad EN alle orders van
    vóór 1-06; reserveringen daar (kolom F) zijn pre-1-06 orders die fysiek
    nog uitgeleverd worden -> die voorraad is NIET vrij voor RugFlow. RugFlow
    maakt alleen NIEUWE orders (ná 1-06). De twee order-sets zijn DISJUNCT,
    dus geen dubbel-aftrekken: baseline = kolom H, en herallocateer_open_orders.py
    trekt daar bovenop alleen RugFlow's eigen nieuwe orders af. Eindresultaat:
    vrij = fysiek - oude verplichtingen - nieuwe RugFlow-orders. (De 2026-06-08
    keuze voor kolom D negeerde de oude verplichtingen -> RugFlow toonde te veel
    vrij = oversold-risico.) Backorder/gereserveerd als baseline op 0.
  - MAATWERK-regels (Karpi-code bevat 'MAATWERK') uitgesloten.
  - Rode regels (rood font) = "niet meer inladen". Karpi markeert deze
    PROGRESSIEF ALFABETISCH per lijst, dus de uitsluitlijst is een UNION:
        exclude = bestaande voorraad_uitsluiten.csv  UNION  rode regels nu.
    Uitgesloten artikelen -> voorraad 0 + (bij --commit) toegevoegd aan de csv.
  - 'vast' in DB maar niet in actieve lijst: voorraad -> 0.
  - Nieuw in lijst maar niet in DB: alleen echte vaste maten met vrije
    voorraad > 0 aanmaken (^[A-Z]{3,4}\\d{2}XX, incl. ...RND). Broadloom/rol
    (geen XX-scheiding) worden gelogd + overgeslagen.
  - Ontbrekende kwaliteiten: elke kwaliteitscode in de actieve lijst zonder
    rij in `kwaliteiten` wordt bij --commit code-only aangemaakt (overige
    velden NULL -> Karpi verrijkt gewicht/omschrijving/collectie later). Zo
    blijft de FK producten.kwaliteit_code geldig en hoeft niets op NULL.
  - Negatieve vrije voorraad -> clampen naar 0.

Bestandsformaat: zowel .xls (klassieke export met rode-font-markering via
xlrd) als .xlsx (nieuwere export via openpyxl). LET OP: de .xlsx-export draagt
doorgaans GEEN font-opmaak -> de rode 'niet inladen'-markering ontbreekt dan.
Dat is veilig door de union-uitsluitlijst (bestaande exclusions blijven staan);
er worden enkel geen NIEUWE rode regels in zo'n ronde gedetecteerd.

Gebruik:
  python update_voorraad.py "..\\Voorraadlijst 01-6-2026.xls"            # DRY-RUN
  python update_voorraad.py "..\\Voorraadlijst 01-6-2026.xls" --commit   # schrijft

Vereist: import/.env met SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY.
"""
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

import xlrd
import pandas as pd
from supabase import create_client

from config import SUPABASE_URL, SUPABASE_KEY, BASE_DIR

UITSLUITEN_CSV = BASE_DIR / "import" / "voorraad_uitsluiten.csv"
RAPPORT_DIR = BASE_DIR / "import" / "rapporten"

# Kolom-indices in de .xls (header op rij index 1, data vanaf rij 2)
COL_ARTNR, COL_KARPI, COL_OMS = 0, 1, 2
COL_FYSIEK = 3   # kolom D 'Voorraad' (FYSIEK) — alleen ter referentie/diff
COL_VOORRAAD = 7  # kolom H 'Vrije voorraad' = baseline voor RugFlow (zie docstring)

_QC_RE = re.compile(r"^([A-Z]{3,4})(\d{2})")
_VAST_RE = re.compile(r"^[A-Z]{3,4}\d{2}XX")
_MAAT_RECHT_RE = re.compile(r"XX(\d{3})(\d{3})$")
_MAAT_ROND_RE = re.compile(r"XX(\d{3})RND$")


def parse_args(argv):
    """-> (Path voorraadlijst, bool commit)."""
    commit = "--commit" in argv
    paden = [a for a in argv[1:] if not a.startswith("--")]
    if not paden:
        print("GEBRUIK: python update_voorraad.py \"<pad naar .xls>\" [--commit]")
        print("Voorbeeld: python update_voorraad.py \"..\\Voorraadlijst 01-6-2026.xls\"")
        sys.exit(1)
    pad = Path(paden[0])
    if not pad.is_absolute():
        pad = (Path.cwd() / pad).resolve()
    if not pad.exists():
        print(f"ERROR: bestand niet gevonden: {pad}")
        sys.exit(1)
    return pad, commit


def is_vaste_maat(karpi_code: str) -> bool:
    return bool(_VAST_RE.match(karpi_code or ""))


def _afgeleide_codes(karpi_code: str):
    m = _QC_RE.match(karpi_code or "")
    if not m:
        return None, None, None
    kwal, kleur = m.group(1), m.group(2)
    return kwal, kleur, f"{kwal}_{kleur}"


def _afmeting(karpi_code: str):
    code = karpi_code or ""
    m = _MAAT_RECHT_RE.search(code)
    if m:
        return int(m.group(1)), int(m.group(2)), None
    m = _MAAT_ROND_RE.search(code)
    if m:
        d = int(m.group(1))
        return d, d, "rond"
    return None, None, None


def _num(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def _maak_rij(artnr, karpi, oms, voorraad, fysiek, is_red):
    artnr = str(artnr).strip()
    karpi = str(karpi).strip()
    if not artnr and not karpi:
        return None
    if artnr.endswith(".0"):
        artnr = artnr[:-2]
    return {
        "artikelnr": artnr,
        "karpi_code": karpi,
        "omschrijving": str(oms).strip(),
        "voorraad": _num(voorraad),  # baseline = kolom H
        "fysiek": _num(fysiek),      # kolom D, referentie
        "is_red": is_red,
        "is_maatwerk": "MAATWERK" in karpi.upper(),
    }


def _lees_xls(pad: Path):
    """Klassieke .xls-export met rode-font-opmaak (xlrd)."""
    wb = xlrd.open_workbook(str(pad), formatting_info=True)
    sh = wb.sheet_by_index(0)

    def is_red(r):
        cell = sh.cell(r, COL_ARTNR)
        xf = wb.xf_list[cell.xf_index]
        font = wb.font_list[xf.font_index]
        return wb.colour_map.get(font.colour_index) == (255, 0, 0)

    rijen = []
    for r in range(2, sh.nrows):
        rij = _maak_rij(
            sh.cell(r, COL_ARTNR).value, sh.cell(r, COL_KARPI).value,
            sh.cell(r, COL_OMS).value, sh.cell(r, COL_VOORRAAD).value,
            sh.cell(r, COL_FYSIEK).value, is_red(r),
        )
        if rij:
            rijen.append(rij)
    return rijen


def _is_rood_font(cell) -> bool:
    """openpyxl: True als de font-kleur een expliciet rood (FF0000) is."""
    col = cell.font.color if cell.font else None
    rgb = getattr(col, "rgb", None) if col else None
    if not isinstance(rgb, str):
        return False
    return rgb.upper().lstrip("F").startswith("FF0000") or rgb.upper() in (
        "FFFF0000", "00FF0000", "FF0000")


def _lees_xlsx(pad: Path):
    """Nieuwere .xlsx-export (openpyxl). LET OP: deze export draagt doorgaans
    GEEN font-opmaak -> rode 'niet inladen'-markering ontbreekt. Dankzij de
    union-uitsluitlijst is dat veilig (bestaande exclusions blijven staan);
    er worden enkel geen NIEUWE rode regels in deze ronde gedetecteerd."""
    import openpyxl
    wb = openpyxl.load_workbook(str(pad), data_only=True)
    sh = wb[wb.sheetnames[0]]
    rijen = []
    # openpyxl is 1-based; header op rij 2 -> data vanaf rij 3
    for row in sh.iter_rows(min_row=3, values_only=False):
        c = {cell.column - 1: cell for cell in row}
        rij = _maak_rij(
            c[COL_ARTNR].value if COL_ARTNR in c else "",
            c[COL_KARPI].value if COL_KARPI in c else "",
            c[COL_OMS].value if COL_OMS in c else "",
            c[COL_VOORRAAD].value if COL_VOORRAAD in c else 0,
            c[COL_FYSIEK].value if COL_FYSIEK in c else 0,
            _is_rood_font(c[COL_ARTNR]) if COL_ARTNR in c else False,
        )
        if rij:
            rijen.append(rij)
    return rijen


def lees_lijst(pad: Path):
    """Lees de voorraadlijst (.xls of .xlsx) -> lijst van rij-dicts incl. is_red."""
    if pad.suffix.lower() == ".xlsx":
        return _lees_xlsx(pad)
    return _lees_xls(pad)


def lees_bestaande_uitsluitlijst():
    """artikelnr -> {artikelnr, karpi_code, omschrijving} uit de skip-lijst."""
    out = {}
    if not UITSLUITEN_CSV.exists():
        return out
    with open(UITSLUITEN_CSV, encoding="utf-8") as f:
        rd = csv.reader(f, delimiter=";")
        next(rd, None)  # header
        for row in rd:
            if not row or not row[0].strip():
                continue
            artnr = row[0].strip()
            out[artnr] = {
                "artikelnr": artnr,
                "karpi_code": (row[1].strip() if len(row) > 1 else ""),
                "omschrijving": (row[2].strip() if len(row) > 2 else ""),
            }
    return out


def laad_db_producten(sb):
    """artikelnr -> (product_type, voorraad) voor alle producten (paginated)."""
    out = {}
    start = 0
    while True:
        r = (sb.table("producten")
             .select("artikelnr,product_type,voorraad")
             .range(start, start + 999).execute())
        if not r.data:
            break
        for x in r.data:
            out[str(x["artikelnr"])] = (x["product_type"], x["voorraad"] or 0)
        if len(r.data) < 1000:
            break
        start += 1000
    return out


def laad_kwaliteiten(sb):
    """set van bestaande kwaliteit-codes (paginated)."""
    codes = set()
    start = 0
    while True:
        kr = sb.table("kwaliteiten").select("code").range(start, start + 999).execute()
        if not kr.data:
            break
        codes.update(str(k["code"]) for k in kr.data)
        if len(kr.data) < 1000:
            break
        start += 1000
    return codes


def laad_rugflow_verzonden_aftrek(sb, vast_set: set) -> dict:
    """Bereken hoeveel stuks al fysiek verscheept zijn via RugFlow-orders
    (bron_systeem != 'oud_systeem') per vaste-maat artikelnr.

    Logica:
      - Verzonden orders     → te_leveren per orderregel (volledig verscheept)
      - Deels verzonden orders → zending_regels.aantal per orderregel (exact
                                 verscheept, want niet alles is al de deur uit)
      - Oud-systeem-orders (bron_systeem='oud_systeem') NIET meenemen: die
        reserveringen zitten al in kolom F van de oude export en zijn dus al
        verrekend in kolom H (vrije voorraad).
      - Maatwerk-regels (is_maatwerk=True) NIET meenemen: geen stuks-voorraad.
      - Admin-pseudo-artikelen (is_pseudo=True, bijv. VERZEND/VORMTOESLAG/
        DROPSHIP-*) NIET meenemen: geen fysieke voorraad.
      - Alleen artikelen met product_type='vast' tellen mee.

    Retourneert dict: artikelnr (str) -> aftrek_stuks (int >= 0).
    """
    aftrek = defaultdict(int)

    # Exclude admin-pseudo artikelen uit de vast_set
    r_ps = sb.table("producten").select("artikelnr").eq("is_pseudo", True).execute()
    pseudo_artnrs = {str(x["artikelnr"]) for x in r_ps.data}
    vast_niet_pseudo = vast_set - pseudo_artnrs

    # --- Deel 1: volledig Verzonden orders (te_leveren = alles verscheept) ---
    # LET OP: .neq("bron_systeem", "oud_systeem") sluit ook NULL-rijen uit
    # (handmatig aangemaakte orders). Gebruik .or() om NULL mee te nemen.
    NIET_OUD_FILTER = "bron_systeem.is.null,bron_systeem.neq.oud_systeem"
    r_v = (sb.table("orders")
           .select("id")
           .eq("status", "Verzonden")
           .or_(NIET_OUD_FILTER)
           .execute())
    verzonden_ids = [x["id"] for x in r_v.data]

    BATCH = 200
    for i in range(0, len(verzonden_ids), BATCH):
        chunk = verzonden_ids[i:i + BATCH]
        start = 0
        while True:
            r = (sb.table("order_regels")
                 .select("artikelnr,te_leveren,is_maatwerk")
                 .in_("order_id", chunk)
                 .eq("is_maatwerk", False)
                 .gt("te_leveren", 0)
                 .range(start, start + 999)
                 .execute())
            if not r.data:
                break
            for x in r.data:
                artnr = str(x.get("artikelnr") or "").strip()
                if artnr in vast_niet_pseudo:
                    aftrek[artnr] += int(x.get("te_leveren") or 0)
            if len(r.data) < 1000:
                break
            start += 1000

    # --- Deel 2: Deels verzonden orders (exacte verscheepte aantallen via
    #     zending_regels — te_leveren op de orderregel bevat nog het totaal) ---
    r_dv = (sb.table("orders")
            .select("id")
            .eq("status", "Deels verzonden")
            .or_(NIET_OUD_FILTER)
            .execute())
    dv_order_ids = [x["id"] for x in r_dv.data]

    if dv_order_ids:
        # Haal de gekoppelde zendingen op (status Onderweg/Klaar voor verzending/
        # Afgeleverd — gereed_op IS NOT NULL = daadwerkelijk gereed gemeld)
        r_zo = (sb.table("zending_orders")
                .select("zending_id")
                .in_("order_id", dv_order_ids)
                .execute())
        dv_zending_ids = list({x["zending_id"] for x in r_zo.data})

        # Filter: alleen zendingen met gereed_op (daadwerkelijk voltooid)
        if dv_zending_ids:
            r_z = (sb.table("zendingen")
                   .select("id")
                   .in_("id", dv_zending_ids)
                   .not_.is_("gereed_op", "null")
                   .execute())
            voltooide_zending_ids = [x["id"] for x in r_z.data]

            if voltooide_zending_ids:
                # Zending_regels → verscheepte aantallen per order_regel_id
                r_zr = (sb.table("zending_regels")
                        .select("order_regel_id,aantal")
                        .in_("zending_id", voltooide_zending_ids)
                        .execute())
                verzonden_per_regel = defaultdict(int)
                for x in r_zr.data:
                    verzonden_per_regel[x["order_regel_id"]] += int(x["aantal"] or 0)

                # Koppel aan artikelnr, filter op vast + niet-maatwerk
                if verzonden_per_regel:
                    r_or = (sb.table("order_regels")
                            .select("id,artikelnr,is_maatwerk")
                            .in_("id", list(verzonden_per_regel.keys()))
                            .eq("is_maatwerk", False)
                            .execute())
                    for reg in r_or.data:
                        artnr = str(reg.get("artikelnr") or "").strip()
                        if artnr in vast_niet_pseudo:
                            aftrek[artnr] += verzonden_per_regel[reg["id"]]

    return dict(aftrek)


def laad_open_orders_meta(sb) -> dict:
    """Dict van open order-ID → orderdatum (created_at) voor FIFO-sortering.
    Uitsluitingen: Verzonden, Geannuleerd, bron_systeem='oud_systeem'.
    NULL bron_systeem (handmatig) wordt WEL meegenomen."""
    NIET_OUD_FILTER = "bron_systeem.is.null,bron_systeem.neq.oud_systeem"
    orders = {}
    start = 0
    while True:
        r = (sb.table("orders").select("id,created_at")
             .not_.in_("status", ["Verzonden", "Geannuleerd"])
             .or_(NIET_OUD_FILTER)
             .range(start, start + 999).execute())
        if not r.data:
            break
        for x in r.data:
            orders[x["id"]] = x.get("created_at") or ""
        if len(r.data) < 1000:
            break
        start += 1000
    return orders


def laad_te_herallocateer_regels(sb, artnrs: set, open_orders_meta: dict) -> list:
    """Alle open order_regel-IDs voor de gegeven artikelen, gesorteerd op orderdatum
    (oudste eerst = FIFO). oud_systeem-orders zijn al uitgesloten via open_orders_meta."""
    if not artnrs or not open_orders_meta:
        return []
    regels = []  # list of (created_at, regel_id)
    artnr_list = list(artnrs)
    BATCH = 200
    for i in range(0, len(artnr_list), BATCH):
        chunk = artnr_list[i:i + BATCH]
        start = 0
        while True:
            r = (sb.table("order_regels").select("id,order_id")
                 .in_("artikelnr", chunk).eq("is_maatwerk", False).gt("te_leveren", 0)
                 .range(start, start + 999).execute())
            if not r.data:
                break
            for x in r.data:
                if x["order_id"] in open_orders_meta:
                    regels.append((open_orders_meta[x["order_id"]], x["id"]))
            if len(r.data) < 1000:
                break
            start += 1000
    regels.sort(key=lambda x: x[0])  # oudste order eerst
    return [rid for _, rid in regels]


def herallocateer_regels(sb, regel_ids: list):
    """Roep herallocateer_orderregel aan voor de gegeven order_regel IDs."""
    ok = fout = 0
    for n, rid in enumerate(regel_ids, 1):
        try:
            sb.rpc("herallocateer_orderregel", {"p_order_regel_id": rid}).execute()
            ok += 1
        except Exception as e:
            fout += 1
            print(f"  FOUT regel {rid}: {e}")
        if n % 100 == 0:
            print(f"  {n}/{len(regel_ids)} verwerkt...")
    print(f"  KLAAR: {ok} regels herallocateerd, {fout} fouten")


def herbereken_alle_intern(sb):
    """Herbereken gereserveerd-cache voor alle artikelen met actieve voorraad-claims."""
    artikelen = set()
    start = 0
    while True:
        r = (sb.table("order_reserveringen").select("fysiek_artikelnr")
             .eq("bron", "voorraad").eq("status", "actief")
             .range(start, start + 999).execute())
        if not r.data:
            break
        artikelen.update(str(x["fysiek_artikelnr"]) for x in r.data if x.get("fysiek_artikelnr"))
        if len(r.data) < 1000:
            break
        start += 1000
    print(f"  artikelen met actieve voorraad-claim: {len(artikelen)}")
    ok = fout = 0
    for n, artnr in enumerate(sorted(artikelen), 1):
        try:
            sb.rpc("herbereken_product_reservering", {"p_artikelnr": artnr}).execute()
            ok += 1
        except Exception as e:
            fout += 1
            print(f"  FOUT artikel {artnr}: {e}")
        if n % 100 == 0:
            print(f"  {n}/{len(artikelen)} verwerkt...")
    print(f"  KLAAR: {ok} artikelen herberekend, {fout} fouten")


def main():
    pad, commit = parse_args(sys.argv)

    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERROR: Supabase URL/Key ontbreekt. Maak import/.env met")
        print("       SUPABASE_URL=... en SUPABASE_SERVICE_ROLE_KEY=...")
        sys.exit(1)
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    stem = re.sub(r"[^A-Za-z0-9_-]+", "_", pad.stem).strip("_")
    rapport = RAPPORT_DIR / f"voorraad_update_{stem}.xlsx"

    print("=" * 64)
    print(f"VOORRAAD-UPDATE  ({'COMMIT' if commit else 'DRY-RUN'})")
    print(f"Bestand: {pad.name}")
    print("=" * 64)

    rijen = lees_lijst(pad)
    print(f"Lijst gelezen: {len(rijen)} data-rijen")

    # --- Uitsluitlijst: UNION van bestaande + nieuwe rode regels ---
    bestaand = lees_bestaande_uitsluitlijst()
    nieuwe_rode = {
        x["artikelnr"]: {
            "artikelnr": x["artikelnr"], "karpi_code": x["karpi_code"],
            "omschrijving": x["omschrijving"],
        } for x in rijen if x["is_red"]
    }
    nieuw_rood_extra = [a for a in nieuwe_rode if a not in bestaand]

    exclude = dict(bestaand)
    exclude.update(nieuwe_rode)
    exclude_artnr = set(exclude)

    print(f"  bestaande uitsluitlijst : {len(bestaand)}")
    print(f"  rood in deze lijst      : {len(nieuwe_rode)}")
    print(f"  nieuw rood toegevoegd   : {len(nieuw_rood_extra)}")
    print(f"  uitsluitlijst NA union  : {len(exclude_artnr)}")

    actief = {}
    for x in rijen:
        if x["artikelnr"] in exclude_artnr or x["is_maatwerk"]:
            continue
        actief.setdefault(x["artikelnr"], x)

    db = laad_db_producten(sb)
    vast = {a for a, (t, _) in db.items() if t == "vast"}

    # --- RugFlow-aftrek: al verscheepte stuks (Verzonden/Deels verzonden,
    #     NIET oud_systeem) per vaste-maat artikelnr ---
    print("\nRugFlow-aftrek ophalen (Verzonden/Deels verzonden, niet-oud_systeem)...")
    aftrek = laad_rugflow_verzonden_aftrek(sb, vast)
    aftrek_totaal = sum(aftrek.values())
    aftrek_artikelen = len(aftrek)
    print(f"  {aftrek_artikelen} artikelen, {aftrek_totaal} stuks aftrekken van baseline")

    updates = []
    op_0_niet_in_lijst = []
    op_0_uitgesloten = []
    for artnr in vast:
        if artnr in exclude_artnr:
            op_0_uitgesloten.append(artnr)
        elif artnr in actief:
            baseline = max(0, actief[artnr]["voorraad"])
            aangepast = max(0, baseline - aftrek.get(artnr, 0))
            updates.append((artnr, aangepast))
        else:
            op_0_niet_in_lijst.append(artnr)

    nieuw_alle = [x for a, x in actief.items() if a not in db]
    nieuw_vast_alle = [x for x in nieuw_alle if is_vaste_maat(x["karpi_code"])]
    nieuw = [x for x in nieuw_vast_alle if x["voorraad"] > 0]
    nieuw_vast_leeg = [x for x in nieuw_vast_alle if x["voorraad"] <= 0]
    nieuw_broadloom = [x for x in nieuw_alle if not is_vaste_maat(x["karpi_code"])]

    # --- Ontbrekende kwaliteiten: codes in de actieve lijst zonder rij in
    #     kwaliteiten. Worden bij --commit code-only aangemaakt (overige velden
    #     NULL -> Karpi verrijkt gewicht/omschrijving/collectie later). ---
    geldige_kwal = laad_kwaliteiten(sb)
    ontbrekende_kwal = {}
    for x in actief.values():
        kwal, _, _ = _afgeleide_codes(x["karpi_code"])
        if kwal and kwal not in geldige_kwal and kwal not in ontbrekende_kwal:
            ontbrekende_kwal[kwal] = x["karpi_code"]

    skip_types = {"staaltje": 0, "rol": 0, "overig": 0}
    for a, (t, _) in db.items():
        if t in skip_types:
            skip_types[t] += 1

    print("\n--- SAMENVATTING ---")
    print(f"  vast geupdatet (uit lijst)        : {len(updates)}")
    print(f"  vast uitgesloten -> 0             : {len(op_0_uitgesloten)}")
    print(f"  vast niet in lijst -> 0           : {len(op_0_niet_in_lijst)}")
    print(f"  nieuw aanmaken (vaste maat, vrd>0): {len(nieuw)}")
    print(f"  nieuw vaste maat 0/neg -> skip    : {len(nieuw_vast_leeg)}")
    print(f"  nieuw broadloom -> overgeslagen   : {len(nieuw_broadloom)}")
    print(f"  ontbrekende kwaliteiten aanmaken  : {len(ontbrekende_kwal)}"
          + (f"  ({', '.join(sorted(ontbrekende_kwal))})" if ontbrekende_kwal else ""))
    print(f"  uitsluitlijst totaal (union)      : {len(exclude_artnr)}")
    print(f"  overgeslagen staaltje             : {skip_types['staaltje']}")
    print(f"  overgeslagen rol                  : {skip_types['rol']}")
    print(f"  overgeslagen overig               : {skip_types['overig']}")
    print(f"\n--- RUGFLOW-AFTREK (Verzonden, niet-oud_systeem) ---")
    print(f"  artikelen met aftrek              : {aftrek_artikelen}")
    print(f"  totaal stuks aftrek               : {aftrek_totaal}")
    # Top-5 ter controle
    top_aftrek = sorted(aftrek.items(), key=lambda x: -x[1])[:5]
    if top_aftrek:
        print(f"  top-5 artikelen:")
        for a, q in top_aftrek:
            baseline = max(0, actief[a]["voorraad"]) if a in actief else 0
            print(f"    {a}: oud-systeem vrij={baseline}  aftrek={q}  "
                  f"-> voorraad={max(0, baseline - q)}")

    # --- Uitsluitlijst wegschrijven (union, gesorteerd) — alleen bij commit ---
    df_uitsluiten = pd.DataFrame(
        sorted(exclude.values(), key=lambda d: d["karpi_code"] or d["artikelnr"])
    )[["artikelnr", "karpi_code", "omschrijving"]]
    if commit:
        df_uitsluiten.to_csv(UITSLUITEN_CSV, sep=";", index=False)
        print(f"\nUitsluitlijst bijgewerkt: {UITSLUITEN_CSV.name} "
              f"({len(df_uitsluiten)} regels)")
    else:
        print(f"\nDRY-RUN: uitsluitlijst NIET overschreven "
              f"(zou {len(df_uitsluiten)} regels worden).")

    # --- Rapport wegschrijven ---
    RAPPORT_DIR.mkdir(exist_ok=True)
    df_nieuw = pd.DataFrame([{
        "artikelnr": x["artikelnr"], "karpi_code": x["karpi_code"],
        "omschrijving": x["omschrijving"], "voorraad": x["voorraad"],
    } for x in nieuw])
    df_broadloom = pd.DataFrame([{
        "artikelnr": x["artikelnr"], "karpi_code": x["karpi_code"],
        "omschrijving": x["omschrijving"], "voorraad_meters": x["voorraad"],
    } for x in nieuw_broadloom])
    df_op0 = pd.DataFrame({"artikelnr": sorted(op_0_niet_in_lijst)})
    df_nieuw_rood = pd.DataFrame([nieuwe_rode[a] for a in nieuw_rood_extra])
    df_kwal = pd.DataFrame(
        [{"kwaliteit_code": c, "voorbeeld_karpi_code": ontbrekende_kwal[c]}
         for c in sorted(ontbrekende_kwal)]
    )
    df_samenvatting = pd.DataFrame([
        {"Categorie": "bestand", "Aantal": pad.name},
        {"Categorie": "vast geupdatet uit lijst", "Aantal": len(updates)},
        {"Categorie": "vast uitgesloten -> 0", "Aantal": len(op_0_uitgesloten)},
        {"Categorie": "vast niet in lijst -> 0", "Aantal": len(op_0_niet_in_lijst)},
        {"Categorie": "nieuw aangemaakt (vaste maat, vrd>0)", "Aantal": len(nieuw)},
        {"Categorie": "nieuw vaste maat 0/neg overgeslagen", "Aantal": len(nieuw_vast_leeg)},
        {"Categorie": "nieuw broadloom overgeslagen", "Aantal": len(nieuw_broadloom)},
        {"Categorie": "ontbrekende kwaliteiten aangemaakt", "Aantal": len(ontbrekende_kwal)},
        {"Categorie": "uitsluitlijst totaal (union)", "Aantal": len(exclude_artnr)},
        {"Categorie": "nieuw rood toegevoegd deze run", "Aantal": len(nieuw_rood_extra)},
        {"Categorie": "overgeslagen staaltje", "Aantal": skip_types["staaltje"]},
        {"Categorie": "overgeslagen rol", "Aantal": skip_types["rol"]},
        {"Categorie": "overgeslagen overig", "Aantal": skip_types["overig"]},
        {"Categorie": "RugFlow aftrek: artikelen (Verzonden)", "Aantal": aftrek_artikelen},
        {"Categorie": "RugFlow aftrek: stuks (Verzonden)", "Aantal": aftrek_totaal},
        {"Categorie": "modus", "Aantal": "COMMIT" if commit else "DRY-RUN"},
    ])
    df_aftrek = pd.DataFrame([
        {
            "artikelnr": a,
            "aftrek_stuks": q,
            "oud_systeem_vrij": max(0, actief[a]["voorraad"]) if a in actief else 0,
            "nieuwe_voorraad": max(0, (max(0, actief[a]["voorraad"]) if a in actief else 0) - q),
        }
        for a, q in sorted(aftrek.items(), key=lambda x: -x[1])
    ])
    with pd.ExcelWriter(rapport, engine="openpyxl") as w:
        df_samenvatting.to_excel(w, sheet_name="Samenvatting", index=False)
        df_aftrek.to_excel(w, sheet_name="RugFlow_aftrek_verzonden", index=False)
        df_nieuw.to_excel(w, sheet_name="Nieuw_vaste_maat", index=False)
        df_broadloom.to_excel(w, sheet_name="Nieuw_broadloom_skip", index=False)
        df_op0.to_excel(w, sheet_name="Op_0_niet_in_lijst", index=False)
        df_nieuw_rood.to_excel(w, sheet_name="Nieuw_rood_deze_run", index=False)
        df_kwal.to_excel(w, sheet_name="Kwaliteiten_aangemaakt", index=False)
        df_uitsluiten.to_excel(w, sheet_name="Uitsluitlijst_union", index=False)
    print(f"Rapport geschreven: {rapport.name}")

    # --- Voorbereiding stap 2/3: welke order_regels worden geraakt? ---
    # (ook in dry-run berekend, zodat de preview informatief is)
    bijgewerkte_artnrs = (
        {a for a, _ in updates} |
        set(op_0_niet_in_lijst) |
        set(op_0_uitgesloten) |
        {x["artikelnr"] for x in nieuw}
    )
    open_orders_meta = laad_open_orders_meta(sb)
    te_herallocateer = laad_te_herallocateer_regels(sb, bijgewerkte_artnrs, open_orders_meta)
    print(f"\n--- STAP 2/3 PREVIEW ---")
    print(f"  open orders (excl. Verzonden/Geannuleerd/oud_systeem): {len(open_orders_meta)}")
    print(f"  order_regels te herallocateer na commit (FIFO)       : {len(te_herallocateer)}")

    if not commit:
        print("\nDRY-RUN: geen DB-wijzigingen. Draai met --commit om te schrijven.")
        return

    # --- COMMIT: wegschrijven naar Supabase ---
    print("\n--- SCHRIJVEN NAAR SUPABASE ---")
    print(f"RugFlow-aftrek toegepast: {aftrek_artikelen} artikelen, {aftrek_totaal} stuks")
    CHUNK = 100
    # updates bevat al de aangepaste waarden (baseline - aftrek, geclamped naar 0)
    per_waarde = defaultdict(list)
    for a, v in updates:
        per_waarde[v].append(a)
    for a in op_0_uitgesloten + op_0_niet_in_lijst:
        per_waarde[0].append(a)

    totaal = sum(len(v) for v in per_waarde.values())
    gedaan = 0
    for v, artnrs in sorted(per_waarde.items()):
        payload = {"voorraad": v, "vrije_voorraad": v,
                   "backorder": 0, "gereserveerd": 0}
        for i in range(0, len(artnrs), CHUNK):
            sb.table("producten").update(payload).in_(
                "artikelnr", artnrs[i:i + CHUNK]).execute()
        gedaan += len(artnrs)
        print(f"  update voorraad={v}: +{len(artnrs)}  ({gedaan}/{totaal})")

    # --- Ontbrekende kwaliteiten aanmaken (code-only) vóór de product-inserts,
    #     zodat de FK producten.kwaliteit_code -> kwaliteiten geldig blijft. ---
    if ontbrekende_kwal:
        kwal_rows = [{"code": c} for c in sorted(ontbrekende_kwal)]
        sb.table("kwaliteiten").insert(kwal_rows).execute()
        geldige_kwal.update(ontbrekende_kwal)
        print(f"  kwaliteiten aangemaakt: {len(kwal_rows)} "
              f"({', '.join(sorted(ontbrekende_kwal))})")

    rec_new = []
    zonder_kwal = []
    for x in nieuw:
        kwal, kleur, zoek = _afgeleide_codes(x["karpi_code"])
        if kwal not in geldige_kwal:
            zonder_kwal.append(x["karpi_code"])
            kwal = None
        lengte, breedte, vorm = _afmeting(x["karpi_code"])
        v_baseline = max(0, x["voorraad"])
        v_aangepast = max(0, v_baseline - aftrek.get(x["artikelnr"], 0))
        rec_new.append({
            "artikelnr": x["artikelnr"], "karpi_code": x["karpi_code"],
            "omschrijving": x["omschrijving"],
            "voorraad": v_aangepast, "vrije_voorraad": v_aangepast,
            "backorder": 0, "gereserveerd": 0,
            "kwaliteit_code": kwal, "kleur_code": kleur, "zoeksleutel": zoek,
            "lengte_cm": lengte, "breedte_cm": breedte,
            "product_type": "vast", "actief": True,
            "vorm": vorm or "rechthoek",
        })
    for i in range(0, len(rec_new), 500):
        sb.table("producten").insert(rec_new[i:i + 500]).execute()
        print(f"  insert nieuw: {min(i + 500, len(rec_new))}/{len(rec_new)}")
    if zonder_kwal:
        print(f"  ({len(zonder_kwal)} nieuw zonder kwaliteit-link: "
              f"{', '.join(sorted(set(_afgeleide_codes(k)[0] for k in zonder_kwal)))})")

    print("\nKLAAR. DB bijgewerkt.")

    # --- STAP 2: herallocateer open orders voor bijgewerkte artikelen ---
    # Zorgt dat inkoop-claims die nu door voorraad gedekt kunnen worden worden
    # omgezet naar voorraad-claims (en omgekeerd bij dalingen).
    print("\n--- STAP 2: HERALLOCATEER OPEN ORDERS ---")
    print(f"  {len(te_herallocateer)} order_regels herberekenen...")
    herallocateer_regels(sb, te_herallocateer)

    # --- STAP 3: herbereken gereserveerd-cache ---
    # Sluitstap: herstelt producten.gereserveerd/vrije_voorraad voor alle
    # artikelen met actieve voorraad-claims (ook die niet door stap 2 geraakt zijn).
    print("\n--- STAP 3: HERBEREKEN GERESERVEERD-CACHE ---")
    herbereken_alle_intern(sb)

    print("\nALLE STAPPEN KLAAR. Voorraad, allocaties en cache zijn up-to-date.")


if __name__ == "__main__":
    main()
