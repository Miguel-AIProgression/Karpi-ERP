"""EENMALIG: reserveer nog-niet-gesneden oud-systeem maatwerk-orders op rollen.

Default = DRY-RUN: leest alles, draait de allocator, schrijft rapport-CSV's naar
import/rapporten/, maar schrijft NIETS naar de database. Gebruik --commit om de
migratie_blokkering-rijen daadwerkelijk weg te schrijven.

Aanroep:
    python reserveer_maatwerk_migratie.py            # dry-run + rapporten
    python reserveer_maatwerk_migratie.py --commit   # schrijf weg

Zie ADR-0028.
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

import openpyxl
from supabase import create_client

from config import SUPABASE_URL, SUPABASE_KEY
from lib.snijlijst_parser import (
    extract_gesneden_uit_rows,
    is_snijden_uit,
    normaliseer_kleur,
    parse_planning_rij,
)
from lib.strip_allocator import Piece, Roll, alloceer

BASE = Path(__file__).parent.parent
PLANNING_FILE = BASE / "Prod planning wk 23-24-25-26  2026 per 03-06-2026.xlsx"
PLANNING_SHEET = "Snijden Karpi op kwal"
VERSIE_GLOB = "Productieplanning Karpi 2026-*.xlsx"
RAPPORT_DIR = Path(__file__).parent / "rapporten"

# Snijplan-statussen die fysiek lengte op een rol verbruiken.
ACTIEVE_SNIJPLAN_STATUS = ("Gepland", "Snijden", "Gesneden")


def bouw_gesneden_set(versie_paths: list[Path]) -> set[tuple[str, str]]:
    """Union van alle (ordernr, rgl) die in ENIGE versie/sheet gesneden zijn."""
    gesneden: set[tuple[str, str]] = set()
    for fn in versie_paths:
        wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
        try:
            for sh in wb.sheetnames:
                rows = list(wb[sh].iter_rows(values_only=True))
                if rows:
                    gesneden |= extract_gesneden_uit_rows(rows)
        finally:
            wb.close()
    return gesneden


def laad_planning() -> list:
    if not PLANNING_FILE.exists():
        sys.exit(f"Planning-bestand niet gevonden: {PLANNING_FILE}")
    wb = openpyxl.load_workbook(PLANNING_FILE, read_only=True, data_only=True)
    try:
        ws = wb[PLANNING_SHEET]
        rows = list(ws.iter_rows(values_only=True))[2:]  # data vanaf rij-index 2
    finally:
        wb.close()
    regels = []
    for r in rows:
        pr = parse_planning_rij(r)
        if pr is not None:
            regels.append(pr)
    return regels


def _fetch_alle(sb, tabel, kolommen, filters=None):
    """Paginerende fetch (Supabase-default limiet is 1000 rijen)."""
    out = []
    start = 0
    page = 1000
    while True:
        q = sb.table(tabel).select(kolommen).range(start, start + page - 1)
        if filters:
            q = filters(q)
        data = q.execute().data or []
        out.extend(data)
        if len(data) < page:
            break
        start += page
    return out


def laad_rollen(sb) -> list[Roll]:
    rows = _fetch_alle(
        sb, "rollen",
        "id, breedte_cm, lengte_cm, kwaliteit_code, kleur_code, status, in_magazijn_sinds",
        lambda q: q.in_("status", ["beschikbaar", "reststuk", "in_snijplan"]),
    )
    # Reeds door snijplannen verbruikte lengte per rol aftrekken (conservatief).
    snij = _fetch_alle(
        sb, "snijplannen",
        "rol_id, lengte_cm, breedte_cm, geroteerd, status",
        lambda q: q.in_("status", list(ACTIEVE_SNIJPLAN_STATUS)).not_.is_("rol_id", "null"),
    )
    verbruikt: dict[int, int] = {}
    for s in snij:
        if s["rol_id"] is None:
            continue
        # Y-as-verbruik = breedte_cm (niet-geroteerd) of lengte_cm (geroteerd).
        y = s["breedte_cm"] if not s.get("geroteerd") else s["lengte_cm"]
        verbruikt[s["rol_id"]] = verbruikt.get(s["rol_id"], 0) + int(y or 0)

    rollen = []
    for r in rows:
        if not r["breedte_cm"] or not r["lengte_cm"]:
            continue  # placeholder-rollen (PH-*) overslaan
        rest = int(r["lengte_cm"]) - verbruikt.get(r["id"], 0)
        if rest <= 0:
            continue
        rollen.append(Roll(
            id=r["id"],
            breedte_cm=int(r["breedte_cm"]),
            lengte_cm=rest,
            kwaliteit=(r["kwaliteit_code"] or "").strip(),
            kleur=normaliseer_kleur(r["kleur_code"]),
            in_magazijn_sinds=r["in_magazijn_sinds"],
        ))
    return rollen


def regels_naar_pieces(regels, gesneden) -> tuple[list[Piece], dict]:
    """Filter planning-regels en zet ze om naar Piece's. Returns (pieces, stats)."""
    # 'actief' telt planning-regels; 'stuks' telt fysieke stukken (regels met
    # Aantal>1 leveren meerdere strips) zodat de operator dit kan kruisvergelijken
    # met 'Gedekt (blokkeringen)' in de dry-run.
    stats = {"totaal": 0, "snijuit": 0, "gesneden": 0, "actief": 0, "stuks": 0}
    pieces = []
    for pr in regels:
        stats["totaal"] += 1
        if is_snijden_uit(pr.opmerking):
            stats["snijuit"] += 1
            continue
        if (pr.oud_ordernr, pr.oud_orderregel) in gesneden:
            stats["gesneden"] += 1
            continue
        stats["actief"] += 1
        stats["stuks"] += pr.aantal
        pieces.append(Piece(
            oud_ordernr=pr.oud_ordernr,
            oud_orderregel=pr.oud_orderregel,
            kwaliteit=pr.kwaliteit,
            kleur=pr.kleur,
            breedte_nodig_cm=pr.breedte_nodig_cm,
            lengte_verbruikt_cm=pr.lengte_verbruikt_cm,
            aantal=pr.aantal,
        ))
    return pieces, stats


def schrijf_rapporten(blok, ongedekt):
    RAPPORT_DIR.mkdir(exist_ok=True)
    with (RAPPORT_DIR / "migratie_gedekt.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["rol_id", "oud_ordernr", "oud_orderregel", "deel_index",
                    "gereserveerde_lengte_cm", "breedte_nodig_cm", "kwaliteit", "kleur"])
        for b in blok:
            w.writerow([b.rol_id, b.oud_ordernr, b.oud_orderregel, b.deel_index,
                        b.gereserveerde_lengte_cm, b.breedte_nodig_cm, b.kwaliteit, b.kleur])
    with (RAPPORT_DIR / "migratie_ongedekt.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["oud_ordernr", "oud_orderregel", "deel_index", "kwaliteit",
                    "kleur", "breedte_nodig_cm", "lengte_verbruikt_cm", "reden"])
        for o in ongedekt:
            w.writerow([o.oud_ordernr, o.oud_orderregel, o.deel_index, o.kwaliteit,
                        o.kleur, o.breedte_nodig_cm, o.lengte_verbruikt_cm, o.reden])


def schrijf_naar_db(sb, blok):
    records = [{
        "rol_id": b.rol_id,
        "gereserveerde_lengte_cm": b.gereserveerde_lengte_cm,
        "breedte_nodig_cm": b.breedte_nodig_cm,
        "oud_ordernr": b.oud_ordernr,
        "oud_orderregel": b.oud_orderregel,
        "deel_index": b.deel_index,
        "kwaliteit_code": b.kwaliteit,
        "kleur_code": b.kleur,
        "status": "actief",
    } for b in blok]
    for i in range(0, len(records), 500):
        sb.table("migratie_blokkering").upsert(
            records[i:i + 500],
            on_conflict="oud_ordernr,oud_orderregel,deel_index",
        ).execute()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true",
                    help="Schrijf migratie_blokkering weg (anders dry-run).")
    args = ap.parse_args()

    versie_paths = sorted(BASE.glob(VERSIE_GLOB))
    if not versie_paths:
        sys.exit(f"Geen versiebestanden gevonden ({VERSIE_GLOB}) in {BASE}")

    print("Gesneden-union bouwen uit", len(versie_paths), "versiebestanden ...")
    gesneden = bouw_gesneden_set(versie_paths)
    print("  unieke gesneden (ordernr,rgl):", len(gesneden))

    regels = laad_planning()
    pieces, stats = regels_naar_pieces(regels, gesneden)
    print("Planning:", stats)

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    rollen = laad_rollen(sb)
    print("Rollen in pool:", len(rollen))

    blok, ongedekt = alloceer(pieces, rollen)
    print("Gedekt (blokkeringen):", len(blok))
    print("Ongedekt (stuks):", len(ongedekt))

    schrijf_rapporten(blok, ongedekt)
    print("Rapporten in:", RAPPORT_DIR)

    if args.commit:
        print("Wegschrijven naar migratie_blokkering ...")
        schrijf_naar_db(sb, blok)
        print("Klaar:", len(blok), "blokkeringen weggeschreven.")
    else:
        print("DRY-RUN — niets weggeschreven. Gebruik --commit om te schrijven.")


if __name__ == "__main__":
    main()
