"""
Import fijnmazige uitwisselgroepen uit Map1.xlsx naar
`kwaliteit_kleur_uitwisselgroepen`.

Kolommen in Map1.xlsx (zonder header):
  A  artikelnr (legacy, bv. VELV10MAATWERK)
  B  collectie_code (3DI / 1VRIJ / 4WOE / 2MA / SNIJ)
  C  variant_nr (1 of 2)
  D  basis_code (groepssleutel, bv. CISC10)

Afleiden:
  kwaliteit_code = artikelnr[0:4]
  kleur_code     = artikelnr[4:6]

Vereist: migration 078 moet zijn toegepast.
"""
from __future__ import annotations

import sys
from collections import defaultdict

import openpyxl
from config import BASE_DIR, SUPABASE_KEY, SUPABASE_URL
from supabase import create_client

MAP1_FILE = BASE_DIR / "Map1.xlsx"

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: Supabase URL/Key niet gevonden. Check import/.env")
    sys.exit(1)

sb = create_client(SUPABASE_URL, SUPABASE_KEY)

print(f"Laden: {MAP1_FILE.name}")
wb = openpyxl.load_workbook(MAP1_FILE, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

by_key: dict[tuple[str, str, int], dict] = {}
skipped = 0

for row in ws.iter_rows(values_only=True):
    if not row or row[0] is None or row[3] is None:
        continue
    artikelnr = str(row[0]).strip()
    collectie = str(row[1]).strip() if row[1] else None
    variant = int(row[2]) if row[2] is not None else 1
    basis = str(row[3]).strip()

    if len(artikelnr) < 6:
        skipped += 1
        continue
    kw = artikelnr[:4].upper()
    kl = artikelnr[4:6]

    key = (kw, kl, variant)
    # Eerste wins; latere duplicaten negeren (PK is (kw,kl,variant))
    if key in by_key:
        continue
    by_key[key] = {
        "kwaliteit_code": kw,
        "kleur_code": kl,
        "variant_nr": variant,
        "basis_code": basis,
        "collectie_code": collectie,
        "bron_artikelnr": artikelnr,
    }

records = list(by_key.values())
print(f"  {len(records)} unieke (kwaliteit,kleur,variant)-rijen ({skipped} overgeslagen)")

# Groep-verdeling tonen
by_basis: dict[tuple[str, int], int] = defaultdict(int)
for r in records:
    by_basis[(r["basis_code"], r["variant_nr"])] += 1
multi = sum(1 for v in by_basis.values() if v > 1)
print(f"  {len(by_basis)} basis-groepen, {multi} met >1 lid")

# Upsert in chunks
print("\nUpserten in kwaliteit_kleur_uitwisselgroepen...")
CHUNK = 500
for i in range(0, len(records), CHUNK):
    chunk = records[i : i + CHUNK]
    sb.table("kwaliteit_kleur_uitwisselgroepen").upsert(
        chunk, on_conflict="kwaliteit_code,kleur_code,variant_nr"
    ).execute()
    print(f"  {i + len(chunk)}/{len(records)}")

# Verificatie
res = sb.table("kwaliteit_kleur_uitwisselgroepen").select("*", count="exact").limit(1).execute()
print(f"\nTotaal rijen in tabel: {res.count}")

print("Klaar.")
